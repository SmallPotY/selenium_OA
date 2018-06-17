from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import openpyxl
from time import sleep


class Exceldb:

    def __init__(self):
        self.wb = openpyxl.load_workbook('item.xlsx')  # 打开excel文件
        self.sheet = self.wb.active  # 获取激活的工作表
        self.max_column = self.sheet.max_column  # 最大列数
        self.max_row = self.sheet.max_row  # 最大行数
        self.row = 2

    def next(self):
        self.item = [i.value for i in self.sheet[self.row]]
        self.row += 1
        return self.item


class Cargo_binding:
    def __init__(self, b_tiem, e_time):
        self.driver = webdriver.Firefox()  # 实例火狐浏览器对象
        self.bdata = b_tiem
        self.edata = e_time

    # 登陆,跳转
    def login(self):
        self.driver.get("http://help.honotop.com/index.html")  # 打开登陆页面
        self.elem = self.driver.find_element_by_xpath('//*[@id="name"]')  # 查找用户输入框
        self.elem.send_keys('***')  # 模拟输入
        self.elem = self.driver.find_element_by_xpath('//*[@id="pass"]')
        self.elem.send_keys('**')
        self.driver.find_element_by_xpath('//*[@id="login"]').click()  # 模拟点击

    # 填写
    def write_store(self):

        self.driver.get("http://help.honotop.com/settlemanager/AutomaticSettle.html?v=516568")  # 跳转到出库结算

        toolbar = self.driver.find_element_by_xpath('/html/body/div[2]/div[1]/div[1]/div/div/button[12]')
        toolbar.click()

        js = "document.getElementById('oneKeyBindBegin').removeAttribute('readonly')"
        self.driver.execute_script(js)

        js = "document.getElementById('oneKeyBindEnd').removeAttribute('readonly')"
        self.driver.execute_script(js)

        oneKeyBindBegin = self.driver.find_element_by_xpath('//*[@id="oneKeyBindBegin"]')
        oneKeyBindBegin.send_keys(self.bdata)

        oneKeyBindEnd = self.driver.find_element_by_xpath('//*[@id="oneKeyBindEnd"]')
        oneKeyBindEnd.send_keys(self.edata)

    def main_loop_store(self, jss, kh, exceldb, i):
        sleep(2)
        oneKeyBindVendor = self.driver.find_element_by_xpath('//*[@id="oneKeyBindVendor"]')
        oneKeyBindVendor.clear()
        oneKeyBindVendor.send_keys(kh)

        btn = self.driver.find_element_by_xpath('//*[@id="oneKeyBindSearch"]')
        btn.click()
        jz = self.driver.find_element_by_xpath(
            '/html/body/div[9]/div/div/div[2]/div[3]/div[1]/div[2]/div[2]/div').is_displayed()

        while True:
            jz = self.driver.find_element_by_xpath(
                '/html/body/div[9]/div/div/div[2]/div[3]/div[1]/div[2]/div[2]/div').is_displayed()

            if jz:
                pass
            else:
                sleep(1)
                hqjg = self.driver.find_element_by_xpath(
                    '/html/body/div[9]/div/div/div[2]/div[3]/div[1]/div[2]/div[2]/table/tbody/tr/td').text
                print('获取结果：', hqjg, end='')

                if hqjg != '没有找到匹配的记录':
                    hqjg = '存在待绑定记录'
                    print(hqjg)
                    sleep(1)
                    btSelectAll = self.driver.find_element_by_xpath(
                        '/html/body/div[9]/div/div/div[2]/div[3]/div[1]/div[2]/div[2]/table/thead/tr/th[1]/div[1]/input')
                    btSelectAll.send_keys(Keys.SPACE)
                    print('选择条目')
                    sleep(1)
                    toOneKeyBind = self.driver.find_element_by_xpath('// *[@id="toOneKeyBind"]')
                    toOneKeyBind.click()
                    print('点击绑定')
                    sleep(5)
                    oneKeyBindSettleNO = Select(self.driver.find_element_by_id('oneKeyBindSettleNO'))
                    oneKeyBindSettleNO.select_by_value(jss)
                    print('输入结算书')
                    sleep(1)
                    btn = self.driver.find_element_by_xpath('//*[@id="submitOneKeyBind"]')
                    btn.click()
                    print('确认绑定')
                    sleep(1)
                    print('等待服务器响应')
                    sleep(8)
                    print('完成绑定')
                    sleep(1)
                    try:
                        al = self.driver.switch_to_alert()  # 获取弹窗
                        al.accept()  # 确认弹窗
                    except:
                        pass

                dyg = 'J' + str(i + 2)
                exceldb.sheet[dyg] = hqjg
                exceldb.wb.save('item.xlsx')
                print('写入记录')
                break




class Storage_binding:
    def __init__(self, b_tiem, e_time):
        self.driver = webdriver.Firefox()  # 实例火狐浏览器对象
        self.bdata = b_tiem
        self.edata = e_time

    # 登陆,跳转
    def login(self):
        self.driver.get("http://help.honotop.com/index.html")  # 打开登陆页面
        self.elem = self.driver.find_element_by_xpath('//*[@id="name"]')  # 查找用户输入框
        self.elem.send_keys('***')  # 模拟输入
        self.elem = self.driver.find_element_by_xpath('//*[@id="pass"]')
        self.elem.send_keys('***')
        self.driver.find_element_by_xpath('//*[@id="login"]').click()  # 模拟点击
        sleep(1)

    # 填写
    def write_store(self):

        self.driver.get("http://help.honotop.com/SettleManager/SettleInventory.html?v=885745)")  # 跳转到仓储结算

        js = "document.getElementById('searchStartDate').removeAttribute('readonly')"
        self.driver.execute_script(js)  # 执行JS代码

        js = "document.getElementById('searchEndDate').removeAttribute('readonly')"
        self.driver.execute_script(js)

        searchStartDate = self.driver.find_element_by_xpath('//*[@id="searchStartDate"]')
        searchStartDate.send_keys(self.bdata)

        searchEndDate = self.driver.find_element_by_xpath('// *[ @ id = "searchEndDate"]')
        searchEndDate.send_keys(self.edata)

        select = Select(self.driver.find_element_by_id('searchSettleStatus'))
        select.select_by_value('已结算')
        select = Select(self.driver.find_element_by_id('searchStatus'))
        select.select_by_value('待绑定')

    def main_loop_store(self, jss, kh, exceldb, i):
        searchVendor = self.driver.find_element_by_xpath('//*[@id="searchVendor"]')  # 填写客户
        searchVendor.clear()
        searchVendor.send_keys(kh)

        search = self.driver.find_element_by_xpath('//*[@id="search"]')
        search.click()  # 点击查询
        sl = 0
        xy = 0
        sleep(1)

        jz = self.driver.find_element_by_xpath('/html/body/div[2]/div[1]/div[2]/div[2]/div').is_displayed()
        al_text = ''
        dd = 5
        while True:
            jz = self.driver.find_element_by_xpath('/html/body/div[2]/div[1]/div[2]/div[2]/div').is_displayed()
            if jz:
                print('等待查询结果...', xy)
                xy += 1
                sleep(1)
            else:
                print('页面加载完成')
                sleep(5)
                hqjg = self.driver.find_element_by_xpath('//*[@id="SettleTable"]/tbody/tr/td').text
                print('获取结果:', hqjg, end="")

                if hqjg != "没有找到匹配的记录":
                    hqjg = self.driver.find_element_by_xpath(
                        '/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/span[1]').text
                    sl = self.driver.find_element_by_xpath(
                        '/html/body/div[2]/div[1]/div[2]/div[2]/table/tbody/tr[1]/td[8]').text
                    hqjg = hqjg

                    print(hqjg)
                    sleep(5)
                    clikc_btn = self.driver.find_element_by_name('btSelectAll')
                    clikc_btn.send_keys(Keys.SPACE)
                    print('全选条目')
                    sleep(5)
                    clikc_btn = self.driver.find_element_by_xpath('//*[@id="toSettle"]')
                    clikc_btn.click()  # 绑定结算书
                    print('绑定结算书号')

                    sleep(5)
                    select = Select(self.driver.find_element_by_id('SettleNO'))
                    select.select_by_value(jss)  # 填写结算书
                    print('填写结算书号:', jss)
                    sleep(3)
                    clikc_btn = self.driver.find_element_by_xpath('//*[@id="SettleOKEx"]')
                    clikc_btn.click()  # 确定
                    print('提交请求,等待服务器响应')


                    while True:
                        try:
                            jzz = self.driver.find_element_by_xpath('//*[@id="loading"]').is_displayed()
                        except:
                            jzz = False

                        if jzz:
                            sleep(1)
                            dd += 1
                            print('等待响应...', dd)
                        else:
                            clikc_btn = self.driver.find_element_by_name('btSelectAll')
                            clikc_btn.send_keys(Keys.SPACE)

                            # print('加载完成')
                            # stop = input('加载完成，按任意键继续')
                            try:
                                print('准备处理弹窗')
                                # sleep(10)
                                al = self.driver.switch_to_alert()  # 获取弹窗
                                al_text = al.text
                                print('弹窗内容', al_text)
                                al.accept()  # 确认弹窗
                            except:
                                # al_text = '未确认弹窗'
                                print('执行完成')
                                sleep(5)
                            finally:
                                break

                # max_row = exceldb.sheet.max_row
                #
                dyg = 'i' + str(i + 2)

                exceldb.sheet[dyg] = hqjg

                exceldb.wb.save('item.xlsx')
                break






if __name__ == '__main__':
    app = Storage_binding('2018-6-1', '2018-6-16')
    app.login()  # 登陆
    app.write_store()

    exceldb = Exceldb()

    for i in range(exceldb.max_row - 1):
        item = exceldb.next()
        if item[8] != '没有找到匹配的记录':
            print('\n***************************')
            print('正在进行:', item[7], '绑定操作')
            try:
                app.main_loop_store(item[3], item[7], exceldb, i)
            except:
                print("出现网络错误。等待中断修复……")
                sleep(15)

    app.driver.close()
