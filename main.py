import json

import openpyxl
import requests
import function


# 获取结算书
def get_item(b_time, e_time):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:59.0) Gecko/20100101 Firefox/59.0',
        'Host': 'help.honotop.com',
        'Referer': 'http://help.honotop.com/SettleManager/SettleBook.html?v=817988',
        'Origin': 'http://help.honotop.com',
        'Cookie': '****',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json'
    }

    data = {
        'args': ['结算书', b_time, e_time],
        'pageIndex': 1,
        'pageSize': 10,
        'where': " 1=1 AND BOOKGROUP='XF'  and book.type=@p0 and book.SettleDate>=@p1 and book.SettleDate<@p2 and book.status<>'取消' "
    }

    url = 'http://help.honotop.com/WCF/B_SettleBookWCF.svc/GetCommonBooks'
    response = requests.post(url=url, data=json.dumps(data), headers=headers)
    request_return = json.loads(response.text)
    rows = request_return['rows']
    row_count = request_return['total']
    wb = openpyxl.Workbook()  # 打开文件
    sheet = wb.active  # 激活sheet表格
    sheet.title = "结算书号"  # 添加sheet表格名称
    sheet.cell(row=1, column=1, value='序号')
    sheet.cell(row=1, column=2, value='供应商名称')
    sheet.cell(row=1, column=3, value='供应商ID')
    sheet.cell(row=1, column=4, value='结算书号')
    sheet.cell(row=1, column=5, value='Type')
    sheet.cell(row=1, column=6, value='合同号')
    sheet.cell(row=1, column=7, value='总金额')
    sheet.cell(row=1, column=8, value='客户代码')
    sheet.cell(row=1, column=9, value='仓储结算')
    sheet.cell(row=1, column=10, value='出库结算')

    ind = 2
    for i in rows[1:]:
        sheet.cell(row=ind, column=1, value=ind - 1)
        sheet.cell(row=ind, column=2, value=i['VendorName'])
        sheet.cell(row=ind, column=3, value=i['VendorID'])
        sheet.cell(row=ind, column=4, value=i['SettleID'])
        sheet.cell(row=ind, column=5, value=i['Type'])
        sheet.cell(row=ind, column=6, value=i['ContractNO'])
        sheet.cell(row=ind, column=7, value=i['TotalAmount'])
        sheet.cell(row=ind, column=8, value=i['VendorID'] + "." + i['VendorName'])
        print(ind - 1, i['VendorName'], i['VendorID'], i['SettleID'], i['Type'], i['ContractNO'], i['TotalAmount'])
        ind += 1
    while True:
        row_count -= 10
        data['pageIndex'] = data['pageIndex'] + 1
        response = requests.post(url=url, data=json.dumps(data), headers=headers)
        request_return = json.loads(response.text)
        rows = request_return['rows']
        for i in rows[1:]:
            sheet.cell(row=ind, column=1, value=ind - 1)
            sheet.cell(row=ind, column=2, value=i['VendorName'])
            sheet.cell(row=ind, column=3, value=i['VendorID'])
            sheet.cell(row=ind, column=4, value=i['SettleID'])
            sheet.cell(row=ind, column=5, value=i['Type'])
            sheet.cell(row=ind, column=6, value=i['ContractNO'])
            sheet.cell(row=ind, column=7, value=i['TotalAmount'])
            sheet.cell(row=ind, column=8, value=i['VendorID'] + "." + i['VendorName'])
            print(ind - 1, i['VendorName'], i['VendorID'], i['SettleID'], i['Type'], i['ContractNO'], i['TotalAmount'])
            ind += 1

        if row_count < 10:
            wb.save('item.xlsx')
            print('\n\n\n执行完毕')
            return


if __name__ == '__main__':
    select = input("选择所要执行的操作：1.获取结算书; 2.绑定出库结算; 3.绑定仓储结算。(输入1、2、3)\n")

    if select == '1':
        b_time = input('输入开始日期：(例：2018-1-1)\n')
        e_time = input('输入结束日期：(例：2018-1-31)\n')

        try:
            print("脚本开始运行……")
            get_item(b_time, e_time)
        except:
            print("执行出错")



    elif select == '2':
        b_time = input('输入开始日期：(例：2018-1-1)\n')
        e_time = input('输入结束日期：(例：2018-1-31)\n')
        app = function.Cargo_binding(b_time, e_time)
        app.login()  # 登陆
        app.write_store()

        exceldb = function.Exceldb()

        for i in range(exceldb.max_row - 1):
            item = exceldb.next()
            if item[9] != '没有找到匹配的记录':
                print('\n***************************')
                print('正在进行:', item[7], '绑定操作')
                try:
                    app.main_loop_store(item[3], item[7], exceldb, i)
                except:
                    print("出现网络错误。等待中断修复……")
                    function.sleep(15)

        app.driver.close()








    elif select == '3':
        b_time = input('输入开始日期：(例：2018-1-1)\n')
        e_time = input('输入结束日期：(例：2018-1-31)\n')



        while True:

            for_key = 1
            app = function.Storage_binding(b_time, e_time)
            app.login()  # 登陆
            app.write_store()

            exceldb = function.Exceldb()

            for i in range(exceldb.max_row - 1):
                item = exceldb.next()
                if item[8] != '没有找到匹配的记录':
                    print('\n***************************')
                    print('正在进行:', item[7], '绑定操作')
                    try:
                        app.main_loop_store(item[3], item[7], exceldb, i)
                    except:
                        print("出现网络错误。等待中断修复……")
                        function.sleep(15)
                        print("\n重启脚本\n")
                        for_key = 0
                        break


            app.driver.close()

            if for_key:
                break
